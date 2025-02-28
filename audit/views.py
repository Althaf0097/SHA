from django.shortcuts import render, redirect, get_object_or_404
from django.core.files.storage import default_storage
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest
from django.contrib.auth.models import User
from django.db.models import Count, Sum, Q
from django.views.generic import ListView
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.contrib.auth.mixins import LoginRequiredMixin
from .models import FieldAudit, District, Patient
from django.utils import timezone
from datetime import datetime, timedelta
from django.db.models import Count, Sum
from django.db.models.functions import TruncMonth
import pandas as pd
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.urls import reverse


def is_admin(user):
    return user.is_superuser

@login_required
def home(request):
    try:
        # Get audit statistics with error handling
        try:
            total_audits = FieldAudit.objects.count()
            recent_audits = FieldAudit.objects.order_by('-visit_date')[:5]
        except Exception as audit_error:
            print(f'Error fetching audit data: {str(audit_error)}')
            total_audits = 0
            recent_audits = []
            messages.error(request, 'Unable to fetch audit data. Please try refreshing the page.')

        # Get patient statistics with error handling
        try:
            documentation_complete = Patient.objects.filter(missing_records=False).count()
            missing_records = Patient.objects.filter(missing_records=True).count()
            compliance_issues = Patient.objects.filter(Q(money_collection=True) | Q(missing_records=True)).count()
            process_gaps = Patient.objects.filter(Q(money_collection=True) | Q(missing_records=True)).count()
            best_practices = Patient.objects.filter(
                Q(missing_records=False) & 
                Q(money_collection=False) 
            ).count()
        except Exception as patient_error:
            print(f'Error fetching patient statistics: {str(patient_error)}')
            documentation_complete = missing_records = compliance_issues = process_gaps = best_practices = 0
            messages.error(request, 'Unable to fetch patient statistics. Please try refreshing the page.')

        context = {
            'total_audits': total_audits,
            'recent_audits': recent_audits,
            'documentation_complete': documentation_complete,
            'missing_records': missing_records,
            'compliance_issues': compliance_issues,
            'process_gaps': process_gaps,
            'best_practices': best_practices,
        }
        return render(request, 'audit/home.html', context)

    except Exception as e:
        print(f'Unexpected error in home view: {str(e)}')
        messages.error(request, 'An unexpected error occurred. Please try again later.')
        context = {
            'total_audits': 0,
            'recent_audits': [],
            'documentation_complete': 0,
            'missing_records': 0,
            'compliance_issues': 0,
            'process_gaps': 0,
            'best_practices': 0,
        }
        return render(request, 'audit/home.html', context)

@login_required
def audit_form(request):
    if request.method == 'POST':
        try:
            # Validate required fields
            required_fields = ['district', 'hospital_id', 'ehcp_name', 'ehcp_type', 'auditor_name', 
                             'designation', 'visit_date', 'visit_time', 'location']
            for field in required_fields:
                if not request.POST.get(field):
                    messages.error(request, f'{field.replace("_", " ").title()} is required.')
                    return redirect('audit_form')

            # Validate district exists
            try:
                district = District.objects.get(name=request.POST.get('district'))
            except District.DoesNotExist:
                messages.error(request, 'Selected district does not exist.')
                return redirect('audit_form')
            
            # Create audit record with new fields
            audit = FieldAudit.objects.create(
                district=district,
                hospital_id=request.POST.get('hospital_id'),
                ehcp_name=request.POST.get('ehcp_name'),
                ehcp_type=request.POST.get('ehcp_type'),
                auditor_name=request.POST.get('auditor_name'),
                designation=request.POST.get('designation'),
                current_location=request.POST.get('location'),
                latitude=request.POST.get('latitude'),
                longitude=request.POST.get('longitude'),
                visit_date=request.POST.get('visit_date'),
                visit_time=request.POST.get('visit_time'),
                ekgp_patients=request.POST.get('kasp_hospital_record', 0),
                pmjay_patients=request.POST.get('kasp_tms_record', 0),
                beneficiaries=request.POST.get('beneficiaries_visited', 0),
                
                # Findings fields
                findings_type=request.POST.get('findings_type'),
                audit_findings_value=request.POST.get('audit_findings_value'),
                finding_type=request.POST.get('finding_type'),
                abuse_type=request.POST.get('abuse_type'),
                oope_type=request.POST.get('oope_type'),
                
                # HNQA fields
                hnqa_value=request.POST.get('hnqa_value'),
                hnqa_type=request.POST.get('hnqa_type'),
                infrastructure_type=request.POST.get('infrastructure_type'),
                hr_type=request.POST.get('hr_type'),
                services_type=request.POST.get('services_type'),
                
                # Fraudulent activities fields
                fraudulent_value=request.POST.get('fraudulent_value'),
                fraudulent_type=request.POST.get('fraudulent_type'),
                
                observations=request.POST.get('audit_observation', ''),
                signature=request.POST.get('signature_data'),
            )
            
            # Handle photo uploads
            if request.FILES.getlist('audit_photos'):
                photo_paths = []
                for photo in request.FILES.getlist('audit_photos'):
                    file_path = f'audit_photos/{audit.id}/{photo.name}'
                    default_storage.save(file_path, photo)
                    photo_paths.append(file_path)
                audit.photos = photo_paths
                audit.save()
            
            messages.success(request, 'Audit record created successfully. Please add patient details.')
            return redirect('add_patient', audit_id=audit.id)
            
        except Exception as e:
            messages.error(request, f'An error occurred while creating the audit record. Please try again or contact support if the issue persists.')
            # Log the actual error for debugging
            print(f'Error in audit_form: {str(e)}')
    
    districts = District.objects.all()
    context = {
        'districts': districts,
    }
    return render(request, 'audit/audit_form.html', context)

@login_required
def add_patient(request, audit_id):
    try:
        audit = get_object_or_404(FieldAudit, id=audit_id)
    except FieldAudit.DoesNotExist:
        messages.error(request, 'Audit record not found.')
        return redirect('home')
    
    if request.method == 'POST':
        try:
            # Validate required fields
            required_fields = ['case_id', 'patient_name', 'admission_date', 'discharge_date', 'package_name']
            for field in required_fields:
                if not request.POST.get(field):
                    messages.error(request, f'{field.replace("_", " ").title()} is required.')
                    return redirect('add_patient', audit_id=audit_id)

            # Validate dates
            try:
                admission_date = datetime.strptime(request.POST.get('admission_date'), '%Y-%m-%d')
                discharge_date = datetime.strptime(request.POST.get('discharge_date'), '%Y-%m-%d')
                if discharge_date < admission_date:
                    messages.error(request, 'Discharge date cannot be earlier than admission date.')
                    return redirect('add_patient', audit_id=audit_id)
            except ValueError:
                messages.error(request, 'Invalid date format. Please use YYYY-MM-DD format.')
                return redirect('add_patient', audit_id=audit_id)

            # Get selected deviations
            deviations = []
            if request.POST.get('money_collection') == 'on':
                deviations.append('money_collection')
            if request.POST.get('package_upcoding') == 'on':
                deviations.append('package_upcoding')
            if request.POST.get('incomplete_records') == 'on':
                deviations.append('incomplete_records')

            # Handle file uploads
            case_file = request.FILES.get('case_file')
            discharge_summary = request.FILES.get('discharge_summary')
            bills_documents = request.FILES.get('bills_documents')

            patient = Patient.objects.create(
                audit=audit,
                case_id=request.POST.get('case_id'),
                patient_name=request.POST.get('patient_name'),
                mobile_number=request.POST.get('mobile_number', ''),
                admission_date=request.POST.get('admission_date'),
                discharge_date=request.POST.get('discharge_date'),
                package_name=request.POST.get('package_name'),
                package_code=request.POST.get('package_code', ''),
                missing_records=request.POST.get('missing_records') == 'on',
                money_collection=request.POST.get('money_collection') == 'on',
                case_summary=request.POST.get('case_summary', ''),
                deviations=deviations,
                total_oope=request.POST.get('total_oope', 0),
                case_file=case_file,
                discharge_summary=discharge_summary,
                bills_documents=bills_documents
            )
            
            if 'add_another' in request.POST:
                messages.success(request, 'Patient added successfully. Add another patient.')
                return redirect('add_patient', audit_id=audit_id)
            else:
                messages.success(request, 'Patient details saved successfully.')
                return redirect('view_records')
        except Exception as e:
            messages.error(request, f'Error adding patient: {str(e)}')
    
    context = {
        'audit': audit,
    }
    return render(request, 'audit/patient_form.html', context)

class AuditListView(LoginRequiredMixin, ListView):
    model = FieldAudit
    template_name = 'audit/audit_list.html'
    context_object_name = 'audits'
    ordering = ['-visit_date']
    paginate_by = 10

@login_required
def patient_detail(request, patient_id):
    patient = get_object_or_404(Patient, id=patient_id)
    return render(request, 'audit/patient_detail.html', {'patient': patient})

@login_required
def edit_patient(request, patient_id):
    try:
        patient = get_object_or_404(Patient, id=patient_id)
    except Patient.DoesNotExist:
        messages.error(request, 'Patient record not found.')
        return redirect('view_records')
    
    if request.method == 'POST':
        try:
            # Validate required fields
            required_fields = ['case_id', 'patient_name', 'admission_date', 'discharge_date', 'package_name']
            for field in required_fields:
                if not request.POST.get(field):
                    messages.error(request, f'{field.replace("_", " ").title()} is required.')
                    return redirect('edit_patient', patient_id=patient_id)

            # Validate dates
            try:
                admission_date = datetime.strptime(request.POST.get('admission_date'), '%Y-%m-%d')
                discharge_date = datetime.strptime(request.POST.get('discharge_date'), '%Y-%m-%d')
                if discharge_date < admission_date:
                    messages.error(request, 'Discharge date cannot be earlier than admission date.')
                    return redirect('edit_patient', patient_id=patient_id)
            except ValueError:
                messages.error(request, 'Invalid date format. Please use YYYY-MM-DD format.')
                return redirect('edit_patient', patient_id=patient_id)

            # Validate mobile number format if provided
            mobile_number = request.POST.get('mobile_number')
            if mobile_number and not mobile_number.isdigit():
                messages.error(request, 'Mobile number should contain only digits.')
                return redirect('edit_patient', patient_id=patient_id)

            # Update patient details
            patient.case_id = request.POST.get('case_id')
            patient.patient_name = request.POST.get('patient_name')
            patient.mobile_number = mobile_number
            patient.admission_date = request.POST.get('admission_date')
            patient.discharge_date = request.POST.get('discharge_date')
            patient.package_name = request.POST.get('package_name')
            patient.package_code = request.POST.get('package_code')
            patient.missing_records = request.POST.get('missing_records') == 'on'
            patient.money_collection = request.POST.get('money_collection') == 'on'
            patient.case_summary = request.POST.get('case_summary')
            patient.total_oope = request.POST.get('total_oope', 0)
            
            # Handle file uploads if present
            if 'patient_photo' in request.FILES:
                if request.FILES['patient_photo'].size > 5242880:  # 5MB limit
                    messages.error(request, 'Patient photo file size should not exceed 5MB.')
                    return redirect('edit_patient', patient_id=patient_id)
                patient.patient_photo = request.FILES['patient_photo']
                
            if 'case_file' in request.FILES:
                if request.FILES['case_file'].size > 10485760:  # 10MB limit
                    messages.error(request, 'Case file size should not exceed 10MB.')
                    return redirect('edit_patient', patient_id=patient_id)
                patient.case_file = request.FILES['case_file']
                
            if 'discharge_summary' in request.FILES:
                if request.FILES['discharge_summary'].size > 10485760:  # 10MB limit
                    messages.error(request, 'Discharge summary file size should not exceed 10MB.')
                    return redirect('edit_patient', patient_id=patient_id)
                patient.discharge_summary = request.FILES['discharge_summary']
                
            patient.save()
            messages.success(request, 'Patient details updated successfully.')
            return redirect('patient_detail', patient_id=patient.id)
            
        except Exception as e:
            messages.error(request, 'An error occurred while updating patient details. Please try again or contact support if the issue persists.')
            print(f'Error in edit_patient: {str(e)}')
            
    return render(request, 'audit/patient_form.html', {'patient': patient, 'is_edit': True})

@login_required
def view_records(request):
    audits = FieldAudit.objects.select_related('district').all().order_by('-visit_date')
    districts = District.objects.all()
    
    # Filter by hospital name
    ehcp_name = request.GET.get('ehcp_name')
    if ehcp_name:
        audits = audits.filter(ehcp_name__icontains=ehcp_name)
    
    # Filter by district
    district = request.GET.get('district')
    if district:
        audits = audits.filter(district__name=district)
    
    # Filter by hospital type
    ehcp_type = request.GET.get('ehcp_type')
    if ehcp_type:
        audits = audits.filter(ehcp_type=ehcp_type)
    
    # Filter by visit date
    visit_date = request.GET.get('visit_date')
    if visit_date:
        audits = audits.filter(visit_date=visit_date)
    
    context = {
        'audits': audits,
        'districts': districts,
        'download_link': request.build_absolute_uri(reverse('download_all_data'))
    }
    
    return render(request, 'audit/view_records.html', context)

@login_required
def download_all_data(request):
    try:
        # Query data with error handling
        try:
            audits = FieldAudit.objects.select_related('district').prefetch_related('patients').all().order_by('-visit_date')
            if not audits.exists():
                messages.warning(request, 'No audit data available to download.')
                return redirect('view_records')
        except Exception as db_error:
            messages.error(request, 'Failed to retrieve audit data. Please try again later.')
            print(f'Database error in download_all_data: {str(db_error)}')
            return redirect('view_records')

        all_data = []
        try:
            for audit in audits:
                audit_data = {
                    'Hospital ID': audit.hospital_id,
                    'Hospital Name': audit.ehcp_name,
                    'District': audit.district.name if audit.district else 'N/A',
                    'Hospital Type': audit.ehcp_type,
                    'Visit Date': audit.visit_date.strftime('%Y-%m-%d') if audit.visit_date else 'N/A',
                    'Visit Time': audit.visit_time.strftime('%H:%M:%S') if audit.visit_time else 'N/A',
                    'Auditor Name': audit.auditor_name,
                    'Designation': audit.designation,
                    'Location': audit.current_location,
                    'EKGP Patients': audit.ekgp_patients or 0,
                    'PMJAY Patients': audit.pmjay_patients or 0,
                    'Total Beneficiaries': audit.beneficiaries or 0,
                    'Findings Type': audit.findings_type or 'N/A',
                    'Audit Findings Value': audit.audit_findings_value or 'N/A',
                    'Finding Type': audit.finding_type or 'N/A',
                    'Abuse Type': audit.abuse_type or 'N/A',
                    'OOPE Type': audit.oope_type or 'N/A',
                    'HNQA Value': audit.hnqa_value or 'N/A',
                    'HNQA Type': audit.hnqa_type or 'N/A',
                    'Infrastructure Type': audit.infrastructure_type or 'N/A',
                    'HR Type': audit.hr_type or 'N/A',
                    'Services Type': audit.services_type or 'N/A',
                    'Fraudulent Value': audit.fraudulent_value or 'N/A',
                    'Fraudulent Type': audit.fraudulent_type or 'N/A',
                    'Observations': audit.observations or ''
                }
                patients = audit.patients.all()
                for patient in patients:
                    try:
                        patient_data = {
                            'Case ID': patient.case_id,
                            'Patient Name': patient.patient_name,
                            'Mobile Number': patient.mobile_number or 'N/A',
                            'Admission Date': patient.admission_date.strftime('%Y-%m-%d') if patient.admission_date else 'N/A',
                            'Discharge Date': patient.discharge_date.strftime('%Y-%m-%d') if patient.discharge_date else 'N/A',
                            'Package Name': patient.package_name,
                            'Package Code': patient.package_code or 'N/A',
                            'Mandatory Records': 'Yes' if patient.missing_records else 'No',
                            'Money Collection': 'Yes' if patient.money_collection else 'No',
                            'Case Summary': patient.case_summary or '',
                            'Deviations': ', '.join(patient.deviations) if patient.deviations else 'None',
                            'Total OOPE': patient.total_oope or 0,
                            'Patient Photo': patient.patient_photo.url if patient.patient_photo else 'No Photo',
                            'Case File': patient.case_file.url if patient.case_file else 'No File',
                            'Discharge Summary': patient.discharge_summary.url if patient.discharge_summary else 'No Summary',
                            'Bills & Documents': patient.bills_documents.url if patient.bills_documents else 'No Documents'
                        }
                        all_data.append({**audit_data, **patient_data})
                    except Exception as patient_error:
                        print(f'Error processing patient {patient.id}: {str(patient_error)}')
                        continue

            # Generate Excel file
            try:
                df = pd.DataFrame(all_data)
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="audit_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
                with pd.ExcelWriter(response, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='All Data')
                return response
            except Exception as excel_error:
                messages.error(request, 'Failed to generate Excel file. Please try again.')
                print(f'Excel generation error: {str(excel_error)}')
                return redirect('view_records')

        except Exception as processing_error:
            messages.error(request, 'An error occurred while processing the data.')
            print(f'Data processing error: {str(processing_error)}')
            return redirect('view_records')

    except Exception as e:
        messages.error(request, 'An unexpected error occurred. Please try again or contact support.')
        print(f'Unexpected error in download_all_data: {str(e)}')
        return redirect('view_records')

@login_required
def download_all_data(request):
    try:
        # Query data with error handling
        try:
            audits = FieldAudit.objects.select_related('district').prefetch_related('patients').all().order_by('-visit_date')
            if not audits.exists():
                messages.warning(request, 'No audit data available to download.')
                return redirect('view_records')
        except Exception as db_error:
            messages.error(request, 'Failed to retrieve audit data. Please try again later.')
            print(f'Database error in download_all_data: {str(db_error)}')
            return redirect('view_records')

        all_data = []
        try:
            for audit in audits:
                audit_data = {
                    'Hospital ID': audit.hospital_id,
                    'Hospital Name': audit.ehcp_name,
                    'District': audit.district.name if audit.district else 'N/A',
                    'Hospital Type': audit.ehcp_type,
                    'Visit Date': audit.visit_date.strftime('%Y-%m-%d') if audit.visit_date else 'N/A',
                    'Visit Time': audit.visit_time.strftime('%H:%M:%S') if audit.visit_time else 'N/A',
                    'Auditor Name': audit.auditor_name,
                    'Designation': audit.designation,
                    'Location': audit.current_location,
                    'EKGP Patients': audit.ekgp_patients or 0,
                    'PMJAY Patients': audit.pmjay_patients or 0,
                    'Total Beneficiaries': audit.beneficiaries or 0,
                    'Findings Type': audit.findings_type or 'N/A',
                    'Audit Findings Value': audit.audit_findings_value or 'N/A',
                    'Finding Type': audit.finding_type or 'N/A',
                    'Abuse Type': audit.abuse_type or 'N/A',
                    'OOPE Type': audit.oope_type or 'N/A',
                    'HNQA Value': audit.hnqa_value or 'N/A',
                    'HNQA Type': audit.hnqa_type or 'N/A',
                    'Infrastructure Type': audit.infrastructure_type or 'N/A',
                    'HR Type': audit.hr_type or 'N/A',
                    'Services Type': audit.services_type or 'N/A',
                    'Fraudulent Value': audit.fraudulent_value or 'N/A',
                    'Fraudulent Type': audit.fraudulent_type or 'N/A',
                    'Observations': audit.observations or ''
                }
                patients = audit.patients.all()
                for patient in patients:
                    try:
                        patient_data = {
                            'Case ID': patient.case_id,
                            'Patient Name': patient.patient_name,
                            'Mobile Number': patient.mobile_number or 'N/A',
                            'Admission Date': patient.admission_date.strftime('%Y-%m-%d') if patient.admission_date else 'N/A',
                            'Discharge Date': patient.discharge_date.strftime('%Y-%m-%d') if patient.discharge_date else 'N/A',
                            'Package Name': patient.package_name,
                            'Package Code': patient.package_code or 'N/A',
                            'Mandatory Records': 'Yes' if patient.missing_records else 'No',
                            'Money Collection': 'Yes' if patient.money_collection else 'No',
                            'Case Summary': patient.case_summary or '',
                            'Deviations': ', '.join(patient.deviations) if patient.deviations else 'None',
                            'Total OOPE': patient.total_oope or 0,
                            'Patient Photo': patient.patient_photo.url if patient.patient_photo else 'No Photo',
                            'Case File': patient.case_file.url if patient.case_file else 'No File',
                            'Discharge Summary': patient.discharge_summary.url if patient.discharge_summary else 'No Summary',
                            'Bills & Documents': patient.bills_documents.url if patient.bills_documents else 'No Documents'
                        }
                        all_data.append({**audit_data, **patient_data})
                    except Exception as patient_error:
                        print(f'Error processing patient {patient.id}: {str(patient_error)}')
                        continue

            # Generate Excel file
            try:
                df = pd.DataFrame(all_data)
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="audit_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
                with pd.ExcelWriter(response, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='All Data')
                return response
            except Exception as excel_error:
                messages.error(request, 'Failed to generate Excel file. Please try again.')
                print(f'Excel generation error: {str(excel_error)}')
                return redirect('view_records')

        except Exception as processing_error:
            messages.error(request, 'An error occurred while processing the data.')
            print(f'Data processing error: {str(processing_error)}')
            return redirect('view_records')

    except Exception as e:
        messages.error(request, 'An unexpected error occurred. Please try again or contact support.')
        print(f'Unexpected error in download_all_data: {str(e)}')
        return redirect('view_records')

@login_required
def report(request):
    # Get overall statistics
    total_audits = FieldAudit.objects.count()
    total_districts = District.objects.count()
    total_patients = Patient.objects.count()
    
    # Get district-wise audit counts
    district_stats = District.objects.annotate(
        audit_count=Count('fieldaudit'),
        total_patients=Sum('fieldaudit__beneficiaries')
    )
    
    # Get monthly audit counts
    current_year = timezone.now().year
    monthly_stats = FieldAudit.objects.filter(
        visit_date__year=current_year
    ).values('visit_date__month').annotate(
        audit_count=Count('id')
    ).order_by('visit_date__month')
    
    context = {
        'total_audits': total_audits,
        'total_districts': total_districts,
        'total_patients': total_patients,
        'district_stats': district_stats,
        'monthly_stats': monthly_stats,
    }
    
    return render(request, 'audit/report.html', context)

@login_required
def hospital_records(request):
    # Get all unique hospitals from audits
    hospitals = FieldAudit.objects.values('hospital_id', 'ehcp_name', 'district__name').distinct()
    
    context = {
        'hospitals': hospitals,
    }
    return render(request, 'audit/hospital_records.html', context)

@login_required
def hospital_patients(request, hospital_id):
    # Get the hospital details
    try:
        hospital = FieldAudit.objects.filter(hospital_id=hospital_id).latest('visit_date')
    except FieldAudit.DoesNotExist:
        hospital = None
    
    if request.method == 'POST':
        try:
            # Get selected deviations
            deviations = []
            if request.POST.get('money_collection') == 'on':
                deviations.append('money_collection')
            if request.POST.get('package_upcoding') == 'on':
                deviations.append('package_upcoding')
            if request.POST.get('incomplete_records') == 'on':
                deviations.append('incomplete_records')

            # Create new patient
            patient = Patient.objects.create(
                audit=hospital,
                case_id=request.POST.get('case_id'),
                patient_name=request.POST.get('patient_name'),
                mobile_number=request.POST.get('mobile_number', ''),
                admission_date=request.POST.get('admission_date'),
                discharge_date=request.POST.get('discharge_date') or None,
                package_name=request.POST.get('package_name'),
                package_code=request.POST.get('package_code', ''),
                missing_records=request.POST.get('missing_records') != 'on',  # Inverted because form shows "Maintained"
                case_summary=request.POST.get('case_summary', ''),
                remarks=request.POST.get('case_summary', ''),  # Copy case_summary to remarks
                deviations=deviations,
                total_oope=request.POST.get('total_oope', 0)
            )
            
            # Handle patient photo
            if 'patient_photo' in request.FILES:
                patient.patient_photo = request.FILES['patient_photo']
                patient.save()

            messages.success(request, 'Patient added successfully.')
            return redirect('hospital_patients', hospital_id=hospital_id)
        except Exception as e:
            messages.error(request, f'Error adding patient: {str(e)}')
    
    # Get all patients for this hospital
    patients = Patient.objects.filter(
        audit__hospital_id=hospital_id
    ).select_related('audit').order_by('-admission_date')
    
    # Get all districts for hospital creation form
    districts = District.objects.all()
    
    context = {
        'hospital': hospital,
        'patients': patients,
        'hospital_id': hospital_id,
        'districts': districts
    }
    return render(request, 'audit/hospital_patients.html', context)

@login_required
def get_patient(request, patient_id):
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
        
    try:
        patient = get_object_or_404(Patient, id=patient_id)
        data = {
            'id': patient.id,
            'case_id': patient.case_id,
            'patient_name': patient.patient_name,
            'mobile_number': patient.mobile_number,
            'admission_date': patient.admission_date.strftime('%Y-%m-%d'),
            'discharge_date': patient.discharge_date.strftime('%Y-%m-%d') if patient.discharge_date else None,
            'package_name': patient.package_name,
            'package_code': patient.package_code,
            'missing_records': patient.missing_records,
            'money_collection': patient.money_collection,
            'case_summary': patient.case_summary,
            'deviations': patient.deviations,
            'total_oope': patient.total_oope,
            'case_file': patient.case_file.url if patient.case_file else None,
            'discharge_summary': patient.discharge_summary.url if patient.discharge_summary else None,
            'bills_documents': patient.bills_documents.url if patient.bills_documents else None
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def edit_patient(request, patient_id):
    patient = get_object_or_404(Patient, id=patient_id)
    
    if request.method == 'POST':
        try:
            # Get selected deviations
            deviations = []
            if request.POST.get('money_collection') == 'on':
                deviations.append('money_collection')
            if request.POST.get('package_upcoding') == 'on':
                deviations.append('package_upcoding')
            if request.POST.get('incomplete_records') == 'on':
                deviations.append('incomplete_records')

            # Update patient fields
            patient.case_id = request.POST.get('case_id')
            patient.patient_name = request.POST.get('patient_name')
            patient.mobile_number = request.POST.get('mobile_number', '')
            patient.admission_date = request.POST.get('admission_date')
            patient.discharge_date = request.POST.get('discharge_date')
            patient.package_name = request.POST.get('package_name')
            patient.package_code = request.POST.get('package_code', '')
            patient.missing_records = request.POST.get('missing_records') == 'on'
            patient.money_collection = request.POST.get('money_collection') == 'on'
            patient.case_summary = request.POST.get('case_summary', '')
            patient.deviations = deviations
            patient.total_oope = request.POST.get('total_oope', 0)

            # Handle file uploads
            if 'case_file' in request.FILES:
                patient.case_file = request.FILES['case_file']
            if 'discharge_summary' in request.FILES:
                patient.discharge_summary = request.FILES['discharge_summary']
            if 'bills_documents' in request.FILES:
                patient.bills_documents = request.FILES['bills_documents']

            patient.save()
            messages.success(request, 'Patient details updated successfully.')
            return redirect('view_records')
        except Exception as e:
            messages.error(request, f'Error updating patient: {str(e)}')
    
    context = {
        'patient': patient,
    }
    return render(request, 'audit/edit_patient.html', context)

@login_required
@require_http_methods(["POST"])
def delete_patient(request, patient_id):
    try:
        patient = get_object_or_404(Patient, id=patient_id)
        patient.delete()
        messages.success(request, 'Patient record deleted successfully.')
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def get_hospital(request, hospital_id):
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
        
    try:
        hospital = FieldAudit.objects.filter(hospital_id=hospital_id).latest('visit_date')
        data = {
            'hospital_id': hospital.hospital_id,
            'ehcp_name': hospital.ehcp_name,
            'ehcp_type': hospital.ehcp_type,
            'district': hospital.district.name,
            'auditor_name': hospital.auditor_name,
            'designation': hospital.designation,
            'current_location': hospital.current_location,
            'latitude': hospital.latitude,
            'longitude': hospital.longitude,
            'visit_date': hospital.visit_date.strftime('%Y-%m-%d'),
            'visit_time': hospital.visit_time.strftime('%H:%M:%S') if hospital.visit_time else '',
            'ekgp_patients': hospital.ekgp_patients,
            'pmjay_patients': hospital.pmjay_patients,
            'beneficiaries': hospital.beneficiaries,
            'findings_type': hospital.findings_type,
            'audit_findings_value': hospital.audit_findings_value,
            'finding_type': hospital.finding_type,
            'abuse_type': hospital.abuse_type,
            'oope_type': hospital.oope_type,
            'hnqa_value': hospital.hnqa_value,
            'hnqa_type': hospital.hnqa_type,
            'infrastructure_type': hospital.infrastructure_type,
            'hr_type': hospital.hr_type,
            'services_type': hospital.services_type,
            'fraudulent_value': hospital.fraudulent_value,
            'fraudulent_type': hospital.fraudulent_type,
            'observations': hospital.observations
        }
        return JsonResponse(data)
    except FieldAudit.DoesNotExist:
        return JsonResponse({'error': 'Hospital not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@require_http_methods(["POST"])
def edit_hospital(request, hospital_id):
    try:
        hospital = get_object_or_404(FieldAudit, hospital_id=hospital_id)
        
        # Get district
        district_name = request.POST.get('district')
        district = get_object_or_404(District, name=district_name)
        
        # Update hospital fields
        hospital.ehcp_name = request.POST.get('ehcp_name')
        hospital.ehcp_type = request.POST.get('ehcp_type')
        hospital.district = district
        hospital.auditor_name = request.POST.get('auditor_name')
        hospital.designation = request.POST.get('designation')
        hospital.current_location = request.POST.get('current_location') or hospital.ehcp_name  # Use hospital name if location not provided
        hospital.latitude = request.POST.get('latitude')
        hospital.longitude = request.POST.get('longitude')
        hospital.visit_date = request.POST.get('visit_date')
        hospital.visit_time = request.POST.get('visit_time')
        hospital.ekgp_patients = request.POST.get('ekgp_patients', 0)
        hospital.pmjay_patients = request.POST.get('pmjay_patients', 0)
        hospital.beneficiaries = request.POST.get('beneficiaries', 0)
        hospital.audit_findings = request.POST.get('audit_findings', 'No')
        hospital.hnqa_related = request.POST.get('hnqa_related', 'No')
        hospital.other_fraudulent_activities = request.POST.get('other_fraudulent_activities', 'No')
        hospital.observations = request.POST.get('observations', '')
        
        hospital.save()
        messages.success(request, 'Hospital details updated successfully.')
        return JsonResponse({'success': True})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@require_http_methods(["POST"])
def delete_hospital(request, hospital_id):
    try:
        hospital = FieldAudit.objects.get(hospital_id=hospital_id)
        hospital.delete()
        messages.success(request, 'Hospital record deleted successfully.')
        return JsonResponse({'success': True})
    except FieldAudit.DoesNotExist:
        return JsonResponse({'error': 'Hospital not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def admin_panel(request):
    if not request.user.is_superuser:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('home')
    
    try:
        # Get statistics
        stats = {
            'total_patients': Patient.objects.count(),
            'total_audits': FieldAudit.objects.count(),
            'total_districts': District.objects.count(),
            'total_users': User.objects.count(),
        }
        
        # Get recent audits
        recent_audits = FieldAudit.objects.select_related('district').order_by('-created_at')[:5]
        
        # Get districts
        districts = District.objects.annotate(
            audit_count=Count('fieldaudit'),
            patient_count=Count('fieldaudit__patients')
        ).order_by('name')
        
        # Get users
        users = User.objects.all().order_by('-date_joined')
        
        context = {
            'stats': stats,
            'recent_audits': recent_audits,
            'districts': districts,
            'users': users,
        }
        
        return render(request, 'audit/admin_panel.html', context)
        
    except Exception as e:
        messages.error(request, f'An error occurred: {str(e)}')
        return render(request, 'audit/admin_panel.html', {'error': str(e)})

@login_required
def add_district(request):
    if not request.user.is_superuser:
        return JsonResponse({'status': 'error', 'message': 'Access denied'})
    
    if request.method == 'POST':
        name = request.POST.get('name')
        if District.objects.filter(name=name).exists():
            return JsonResponse({'status': 'error', 'message': 'District already exists'})
        
        district = District.objects.create(name=name)
        return JsonResponse({
            'status': 'success',
            'message': 'District added successfully',
            'district': {'id': district.id, 'name': district.name}
        })
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

@login_required
def edit_district(request, district_id):
    if not request.user.is_superuser:
        return JsonResponse({'status': 'error', 'message': 'Access denied'})
    
    try:
        district = District.objects.get(id=district_id)
        if request.method == 'POST':
            name = request.POST.get('name')
            if District.objects.filter(name=name).exclude(id=district_id).exists():
                return JsonResponse({'status': 'error', 'message': 'District name already exists'})
            
            district.name = name
            district.save()
            return JsonResponse({
                'status': 'success',
                'message': 'District updated successfully',
                'district': {'id': district.id, 'name': district.name}
            })
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
    except District.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'District not found'})

@login_required
def delete_district(request, district_id):
    if not request.user.is_superuser:
        return JsonResponse({'status': 'error', 'message': 'Access denied'})
    
    try:
        district = District.objects.get(id=district_id)
        if FieldAudit.objects.filter(district=district).exists():
            return JsonResponse({'status': 'error', 'message': 'Cannot delete district with associated audits'})
        
        district.delete()
        return JsonResponse({'status': 'success', 'message': 'District deleted successfully'})
    except District.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'District not found'})

@login_required
def add_user(request):
    if not request.user.is_superuser:
        return JsonResponse({'status': 'error', 'message': 'Access denied'})
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        is_staff = request.POST.get('is_staff') == 'true'
        
        if User.objects.filter(username=username).exists():
            return JsonResponse({'status': 'error', 'message': 'Username already exists'})
        
        user = User.objects.create_user(
            username=username,
            password=password,
            is_staff=is_staff
        )
        return JsonResponse({
            'status': 'success',
            'message': 'User added successfully',
            'user': {'id': user.id, 'username': user.username, 'is_staff': user.is_staff}
        })
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

@login_required
def edit_user(request, user_id):
    if not request.user.is_superuser:
        return JsonResponse({'status': 'error', 'message': 'Access denied'})
    
    try:
        user = User.objects.get(id=user_id)
        if request.method == 'POST':
            username = request.POST.get('username')
            password = request.POST.get('password')
            is_staff = request.POST.get('is_staff') == 'true'
            
            if User.objects.filter(username=username).exclude(id=user_id).exists():
                return JsonResponse({'status': 'error', 'message': 'Username already exists'})
            
            user.username = username
            if password:
                user.set_password(password)
            user.is_staff = is_staff
            user.save()
            
            return JsonResponse({
                'status': 'success',
                'message': 'User updated successfully',
                'user': {'id': user.id, 'username': user.username, 'is_staff': user.is_staff}
            })
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
    except User.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'User not found'})

@login_required
def delete_user(request, user_id):
    if not request.user.is_superuser:
        return JsonResponse({'status': 'error', 'message': 'Access denied'})
    
    try:
        if int(user_id) == request.user.id:
            return JsonResponse({'status': 'error', 'message': 'Cannot delete your own account'})
        
        user = User.objects.get(id=user_id)
        user.delete()
        return JsonResponse({'status': 'success', 'message': 'User deleted successfully'})
    except User.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'User not found'})

@login_required
def audit_statistics(request):
    if not request.user.is_superuser:
        return JsonResponse({'status': 'error', 'message': 'Access denied'})
    
    total_audits = FieldAudit.objects.count()
    completed_audits = FieldAudit.objects.filter(status='Completed').count()
    pending_audits = FieldAudit.objects.filter(status='Pending').count()
    in_progress_audits = FieldAudit.objects.filter(status='In Progress').count()
    
    fraud_cases = Patient.objects.filter(
        Q(money_collection=True) | 
        Q(missing_records=True)
    ).count()
    
    district_stats = District.objects.annotate(
        audit_count=Count('fieldaudit'),
        fraud_count=Count('fieldaudit__patients', filter=Q(
            fieldaudit__patients__money_collection=True) | 
            Q(fieldaudit__patients__missing_records=True)
        )
    ).values('name', 'audit_count', 'fraud_count')
    
    return JsonResponse({
        'status': 'success',
        'data': {
            'total_audits': total_audits,
            'completed_audits': completed_audits,
            'pending_audits': pending_audits,
            'in_progress_audits': in_progress_audits,
            'fraud_cases': fraud_cases,
            'district_stats': list(district_stats)
        }
    })

@login_required
def admin_dashboard(request):
    if not request.user.is_superuser:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('home')

    # Get current date
    current_date = timezone.now()
    
    # Basic statistics
    total_patients = Patient.objects.count()
    total_audits = FieldAudit.objects.count()
    
    # Mock data for demonstration
    todays_appointments = 15
    available_doctors = 8
    total_doctors = 10
    monthly_revenue = "45,000"
    
    # Get recent audits
    recent_audits = FieldAudit.objects.order_by('-created_at')[:5]
    
    # Get monthly audit counts
    six_months_ago = current_date - timedelta(days=180)
    monthly_audits = FieldAudit.objects.filter(
        created_at__gte=six_months_ago
    ).annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(
        count=Count('id')
    ).order_by('month')
    
    # Calculate compliance rate
    total_reviews = Patient.objects.count()
    compliant = Patient.objects.filter(
        Q(money_collection=False) & 
        Q(missing_records=False)
    ).count()
    compliance_rate = (compliant / total_reviews * 100) if total_reviews > 0 else 0
    
    context = {
        'current_date': current_date,
        'total_patients': total_patients,
        'todays_appointments': todays_appointments,
        'available_doctors': available_doctors,
        'total_doctors': total_doctors,
        'monthly_revenue': monthly_revenue,
        'recent_audits': recent_audits,
        'monthly_audits': list(monthly_audits),
        'compliance_rate': compliance_rate,
    }
    
    return render(request, 'audit/admin_dashboard.html', context)

@login_required
def create_hospital(request):
    if request.method == 'POST':
        try:
            district = get_object_or_404(District, id=request.POST.get('district'))
            hospital = FieldAudit.objects.create(
                district=district,
                hospital_id=request.POST.get('hospital_id'),
                ehcp_name=request.POST.get('ehcp_name'),
                ehcp_type=request.POST.get('ehcp_type'),
                auditor_name=request.user.get_full_name() or request.user.username,
                designation='Auditor',  # Default value
                current_location=district.name,
                latitude=request.POST.get('latitude'),
                longitude=request.POST.get('longitude'),
                visit_date=timezone.now().date(),
                visit_time=timezone.now().time(),
            )
            messages.success(request, 'Hospital created successfully.')
            return redirect('hospital_patients', hospital_id=hospital.hospital_id)
        except Exception as e:
            messages.error(request, f'Error creating hospital: {str(e)}')
            return redirect('hospital_patients', hospital_id=request.POST.get('hospital_id'))
    
    return redirect('hospital_records')

@login_required
def download_patient_data(request, hospital_id):
    hospital = get_object_or_404(FieldAudit, hospital_id=hospital_id)
    patients = Patient.objects.filter(audit__hospital_id=hospital_id)
    
    # Create data for Excel
    data = []
    for patient in patients:
        # Get deviation list as comma-separated string
        deviations = ', '.join(patient.deviations) if patient.deviations else 'None'
        
        # Get document status
        documents = []
        if patient.patient_photo:
            documents.append('Photo')
        if patient.case_file:
            documents.append('Case File')
        if patient.discharge_summary:
            documents.append('Discharge Summary')
        if patient.bills_documents:
            documents.append('Bills')
        documents_status = ', '.join(documents) if documents else 'None'
        
        data.append({
            'Hospital Name': hospital.ehcp_name,
            'Hospital ID': hospital.hospital_id,
            'District': hospital.district.name,
            'Case ID': patient.case_id,
            'Patient Name': patient.patient_name,
            'Mobile Number': patient.mobile_number,
            'Admission Date': patient.admission_date.strftime('%d/%m/%Y') if patient.admission_date else '',
            'Discharge Date': patient.discharge_date.strftime('%d/%m/%Y') if patient.discharge_date else '',
            'Package Name': patient.package_name,
            'Package Code': patient.package_code,
            'Mandatory Records': 'No' if patient.missing_records else 'Yes',
            'Deviations': deviations,
            'OOPE Amount': patient.total_oope,
            'Documents': documents_status
        })
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Create Excel response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{hospital.ehcp_name}_patients_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    # Write to Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Patient Data')
        
        # Auto-adjust columns' width
        worksheet = writer.sheets['Patient Data']
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            ) + 2
            worksheet.column_dimensions[chr(65 + idx)].width = min(max_length, 50)
    
    return response

@login_required
def download_patient_excel(request, hospital_id):
    hospital = get_object_or_404(FieldAudit, hospital_id=hospital_id)
    patients = Patient.objects.filter(audit__hospital_id=hospital_id)
    
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Patient Details"
    
    # Define styles
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    hospital_font = Font(name='Arial', size=11, bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add hospital details in first row
    ws.merge_cells('A1:B1')
    ws.merge_cells('C1:D1')
    ws.merge_cells('E1:F1')
    
    ws['A1'] = f"Hospital Name: {hospital.ehcp_name}"
    ws['C1'] = f"Hospital ID: {hospital.hospital_id}"
    ws['E1'] = f"District: {hospital.district.name if hospital.district else ''}"
    
    # Style hospital details
    for cell in ['A1', 'C1', 'E1']:
        ws[cell].font = hospital_font
        ws[cell].border = border
        ws[cell].alignment = Alignment(horizontal='left', vertical='center')
    
    # Add headers in second row
    headers = [
        'SI.NO', 'Case ID', 'Patient Name', 'Mobile Number', 'Admission Date', 
        'Discharge Date', 'Package Name', 'Package Code', 'Mandatory Records',
        'Deviations', 'OOPE Amount'
    ]
    
    # Write and style headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data starting from row 3
    row = 3
    for index, patient in enumerate(patients, 1):
        # Format deviations - convert list to string
        deviations = patient.deviations
        if isinstance(deviations, list):
            deviations = ', '.join(deviations) if deviations else 'None'
        elif deviations is None:
            deviations = 'None'
        
        # Add data
        data = [
            index,  # SI.NO
            patient.case_id,
            patient.patient_name,
            patient.mobile_number,
            patient.admission_date.strftime('%d/%m/%Y') if patient.admission_date else '',
            patient.discharge_date.strftime('%d/%m/%Y') if patient.discharge_date else '',
            patient.package_name,
            patient.package_code,
            'Yes' if patient.missing_records else 'No',
            deviations,
            patient.total_oope if patient.total_oope else 0
        ]
        
        # Write and style each cell
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.font = Font(name='Arial', size=10)
            
            # Center align specific columns
            if col in [1, 3, 4, 5, 8]:  # SI.NO, Mobile, Dates, and Mandatory Records
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Right align amount
            if col == 11:  # OOPE Amount (now column 11 due to SI.NO)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0.00'
        
        row += 1
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column = get_column_letter(col)
        max_length = 0
        column_cells = ws[column]
        
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 25)  # Reduced max width to make columns more compact
    
    # Set row heights
    ws.row_dimensions[1].height = 20  # Hospital details row
    ws.row_dimensions[2].height = 20  # Headers row
    
    # Freeze the header rows
    ws.freeze_panes = 'A3'
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=patient_details_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response
